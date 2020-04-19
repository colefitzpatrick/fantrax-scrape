from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup
import re
import pandas as pd
import os
import time
import openpyxl
os.chdir('c:\\Python\\colefitzpatrick_python')

url = "https://www.fantrax.com/login"

wb_write = openpyxl.load_workbook('fantrax.xlsx')
ws_write = wb_write["Sheet1"]

# create a new Firefox session
driver = webdriver.Firefox()
driver.implicitly_wait(8)
driver.get(url)

#enters login information
username = driver.find_element_by_id("mat-input-0")
username.clear()
username.send_keys("[INSERT USERNAME HERE]")            #<-------username input

password = driver.find_element_by_id("mat-input-1")
password.clear()
password.send_keys("[INSERT PASSWORD HERE]")       #<-------password input

driver.implicitly_wait(3)

#clicks submit
submitbutton = driver.find_element_by_xpath('/html/body/app-root/div/div[2]/div/app-login/div/section/form/div[2]/button/span')
driver.execute_script("arguments[0].click();", submitbutton)

time.sleep(5)

#clicks the league dropdown
dropdown = driver.find_element_by_xpath('/html/body/app-root/div/div[1]/navbar/nav/div/div/section[1]/div[1]/b')
driver.execute_script("arguments[0].click();", dropdown)

driver.implicitly_wait(1)

#selects the appropriate league
myleague = driver.find_element_by_xpath('/html/body/app-root/div/div[1]/navbar/nav/div/div/section[1]/div[1]/div/div/div/league-nav/div[2]/a')
driver.execute_script("arguments[0].click();", myleague)

driver.implicitly_wait(3)

#xpath components that do not change for the standings table
path1 = '/html/body/app-root/div/div[2]/div/app-league-home/div/section/div/div[1]/league-home-standings/pane/section/div[2]/div/league-home-standings-content/table/tbody['
path2 = ']/tr['
path_team = ']/td[2]/a'
path_record = ']/td[3]/span'

existingrecords = {}

for tbody in range(1,4):
    for trrecord in range(2,6):
        teamrecord = driver.find_element_by_xpath(path1 + str(tbody) + path2 + str(trrecord) + path_record).text   #scrapes the full string of the team's record (ex. 51-23-2)
        teamname = driver.find_element_by_xpath(path1 + str(tbody) + path2 + str(trrecord) + path_team).text     #scrapes the team name
        dashpos = teamrecord.find('-')    #finds the dash position
        lenrecord = len(teamrecord)     #gets the length of the record value
        ties = lenrecord - 1      #gets the position of the last dash
        print(teamname)
        print("Ties: " + teamrecord[ties:])
        print("Wins: " + teamrecord[:dashpos])
        equivwins = int(teamrecord[:dashpos]) + (0.5 * int(teamrecord[ties:]))      #wins + 1/2 * ties = equivalent wins
        existingrecords.update( { teamname : equivwins} )    #updates the dictionary with the team/record pair
splitrecord = teamrecord.split('-')   #splits the record string using dash delimiter
totalgamesplayed = int(splitrecord[0]) + int(splitrecord[1]) + int(splitrecord[2])      #sums wins losses and ties to get number of games played
print("Games Played: " + str(totalgamesplayed))
print(existingrecords)
time.sleep(5)

#clicks the players page
players = driver.find_element_by_xpath('/html/body/app-root/div/div[2]/app-leagues-header/nav/div/div[4]/a')
driver.execute_script("arguments[0].click();", players)

time.sleep(5)

#click the status/team dropdown
status_selector = driver.find_element_by_xpath('/html/body/app-root/div/div[2]/div/app-league-players/div/section/filter-panel/div/div[4]/div[1]/mat-form-field/div/div[1]/div[3]/mat-select/div/div[1]')
driver.execute_script("arguments[0].click();", status_selector)

driver.implicitly_wait(1)

#select All Taken Players
try:
    taken_selector = driver.find_element_by_xpath('/html/body/div[4]/div[2]/div/div/div/mat-option[5]/span')
except NoSuchElementException:
    taken_selector = driver.find_element_by_xpath('/html/body/div[5]/div[2]/div/div/div/mat-option[5]/span')
driver.execute_script("arguments[0].click();", taken_selector)

time.sleep(3)  #gives time to load

#select the rows per page dropdown
perpage = driver.find_element_by_xpath('/html/body/app-root/div/div[2]/div/app-league-players/div/section/div[2]/pagination/div[4]/button/span')
driver.execute_script("arguments[0].click();", perpage)

driver.implicitly_wait(1)

#select 500 per page
try:
    fivehundredper = driver.find_element_by_xpath('/html/body/div[4]/div[2]/div/div/div/div/button[5]')
except NoSuchElementException:
    fivehundredper = driver.find_element_by_xpath('/html/body/div[5]/div[2]/div/div/div/div/button[5]')
driver.execute_script("arguments[0].click();", fivehundredper)

time.sleep(3) #gives time to load

soup_level2=BeautifulSoup(driver.page_source, 'lxml')

writerow = 1
for tr in soup_level2.findAll("td", {"class": "ng-star-inserted"}):
    tds = tr.findAll("div", {"class": "scorer__info"})
    for td in tds:
        playername = td.findAll("div", {"class": "scorer__info__name"})
        playerpos = td.findAll("div", {"class": "scorer__info__positions"})
        for player in playername:
            #print(player.text)
            ws_write.cell(row=writerow, column=1).value = player.text     #writes the player name to the first column
        for player1 in playerpos:
            posandteam = player1.findAll("span")
            for entry in posandteam:
                if len(entry.text) > 5 and entry.text.count(',') == 0:
                    continue
                elif entry.text in ['C','1B','2B','3B','OF','SP','SS','RP','P','UT'] or entry.text.count(',') >= 1:
                #    print("Pos: " + entry.text)
                    ws_write.cell(row=writerow, column=2).value = entry.text        #writes the player position(s) to the 2nd column
                elif entry.text == "(R)":
                    continue
                elif entry.text == '-':
                    continue
                elif entry.text == "":
                    continue
                else:
               #     print("Team: " + entry.text[1:])
              #      print("----")
                    ws_write.cell(row=writerow, column=3).value = entry.text[1:]     #writes the player's MLB team to the 3rd column
        writerow += 1

writerow = 1
for tr1 in soup_level2.findAll("tr", {"class": "ng-star-inserted"}):
    tds1 = tr1.findAll("table-cell", {"class": "ng-star-inserted"})
    rowvalues = []
    for td1 in tds1:
        rowvalues.append(td1.text)
    ws_write.cell(row=writerow, column=4).value = rowvalues[1]       #writes the player's fantasy team
    ws_write.cell(row=writerow, column=5).value = rowvalues[5]       #writes the player's total points
    ws_write.cell(row=writerow, column=6).value = rowvalues[6]       #writes the player's points per game
    writerow += 1


wb_write.save('fantrax.xlsx')

